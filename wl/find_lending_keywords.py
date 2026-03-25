# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx')
ws = wb.active

# Search for more specific lending-related keywords
keywords = ['借款', '借钱', '还款', '归还', '欠款', '个人借', '公司借', '借出', '借入', '贷款', '暂借', '周转']
results = []

for row_idx in range(3, ws.max_row + 1):
    row_text = ''
    for col_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            row_text += str(cell.value) + ' '
    
    if any(kw in row_text for kw in keywords):
        date = ws.cell(row=row_idx, column=11).value
        party = ws.cell(row=row_idx, column=7).value
        amount = ws.cell(row=row_idx, column=9).value
        summary = ws.cell(row=row_idx, column=14).value
        results.append({
            'date': date,
            'party': party,
            'amount': amount,
            'summary': summary
        })

# Write to file
with open(r'C:\Users\y2k1\.openclaw\workspace\wl\lending_detailed.txt', 'w', encoding='utf-8') as f:
    f.write(f'Found {len(results)} lending-related records:\n\n')
    for r in results:
        f.write(f"日期: {r['date']}\n")
        f.write(f"交易对手: {r['party']}\n")
        f.write(f"金额: {r['amount']}\n")
        f.write(f"摘要: {r['summary']}\n")
        f.write('-' * 30 + '\n')

print(f'Total: {len(results)} records')

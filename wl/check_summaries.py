# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx', data_only=True)
ws = wb.active

# Search all summary column for any text
# First, let's see all unique summaries
summaries = set()
for row_idx in range(3, min(ws.max_row + 1, 500)):
    summary_cell = ws.cell(row=row_idx, column=14)
    if summary_cell.value:
        summaries.add(str(summary_cell.value)[:20])

print('Sample summaries from the file:')
for s in sorted(summaries)[:30]:
    print(repr(s))

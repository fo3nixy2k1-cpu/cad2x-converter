# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx', data_only=True)
ws = wb.active

# Find rows with 借钱 or 借款
keywords = ['借钱', '借款']
results = []

for row_idx in range(3, ws.max_row + 1):
    summary_cell = ws.cell(row=row_idx, column=14)  # Column 14 is summary
    summary = summary_cell.value
    
    if summary and any(kw in str(summary) for kw in keywords):
        date = ws.cell(row=row_idx, column=11).value
        party = ws.cell(row=row_idx, column=7).value
        amount = ws.cell(row=row_idx, column=9).value
        
        results.append({
            'date': str(date)[:10] if date else '',
            'party': party if party else '',
            'amount': amount if amount else '',
            'summary': summary
        })

print(f'Found {len(results)} records with 借钱/借款:\n')
for r in results:
    print(f"日期: {r['date']}")
    print(f"交易对手: {r['party']}")
    print(f"金额: {r['amount']}")
    print(f"摘要: {r['summary']}")
    print('-' * 30)

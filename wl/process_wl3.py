# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import csv
import sys

# Set output encoding
sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx')
ws = wb.active

# Find ALL rows with '浩之源' in column 7 (counterparty name)
# Read cells as-is and try to match
hanzhiyuan = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[6]:  # Column 7 - counterparty
        # The cell contains the company name - check if it matches
        cell_val = str(row[6])
        
        # We need to check for the specific company - let's check bytes
        try:
            # Encode and decode to check for specific characters
            if b'\xe6\xbf\x95' in cell_val.encode('utf-8') and b'\xe4\xb9\x8b' in cell_val.encode('utf-8') and b'\xe6\xba\x90' in cell_val.encode('utf-8'):
                # This is 浩之源 - now check if it's the full name
                if b'\xe9\x80\x9a\xe4\xbf\xa1' in cell_val.encode('utf-8'):
                    # This is 通信 - so it's 河南浩之源通信工程有限公司
                    date = str(row[10])[:10] if row[10] else ''
                    amount = row[8]
                    counterparty = row[6]
                    hanzhiyuan.append({
                        '日期': date,
                        '交易对手': counterparty,
                        '交易金额': amount
                    })
                    print(f"Found: Date={date}, Amount={amount}")
        except:
            pass

print(f'\nTotal: {len(hanzhiyuan)} rows with 河南浩之源通信工程有限公司')

# Read CMB CSV for 陈亚
chenya = []
with open(r'C:\Users\y2k1\.openclaw\workspace\wl\cmb.csv', 'r', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    for row in reader:
        line = ','.join(row)
        if '陈亚' in line:
            parts = line.split(',')
            if len(parts) >= 7:
                date = parts[0].strip().replace('\t', '')
                income = parts[2].strip().replace('\t', '') if len(parts) > 2 else ''
                outcome = parts[3].strip().replace('\t', '') if len(parts) > 3 else ''
                amount = income if income else ('-' + outcome if outcome else '')
                chenya.append({
                    '日期': date,
                    '交易对手': '陈亚',
                    '交易金额': amount
                })

print(f'Found {len(chenya)} rows for 陈亚')

# Create output
output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = '往来款'

output_ws['A1'] = '日期'
output_ws['B1'] = '交易对手'
output_ws['C1'] = '交易金额'

row_num = 2
for record in chenya:
    output_ws[f'A{row_num}'] = record['日期']
    output_ws[f'B{row_num}'] = record['交易对手']
    output_ws[f'C{row_num}'] = record['交易金额']
    row_num += 1

for record in hanzhiyuan:
    output_ws[f'A{row_num}'] = record['日期']
    output_ws[f'B{row_num}'] = record['交易对手']
    output_ws[f'C{row_num}'] = record['交易金额']
    row_num += 1

output_wb.save(r'C:\Users\y2k1\.openclaw\workspace\wl\往来款汇总.xlsx')
print(f'\nSaved! Total: {row_num-1} rows')

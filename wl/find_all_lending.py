# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx')
ws = wb.active

# Find all expense records (negative amounts) and group by counterparty
# Also search for keywords that might indicate lending
lending_keywords = ['借', '贷', '还款', '归还', '欠', '个人借', '公司借']
results = []

for row_idx in range(3, ws.max_row + 1):
    row_text = ''
    for col_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            row_text += str(cell.value) + ' '
    
    # Check if any lending keyword exists
    has_keyword = any(kw in row_text for kw in lending_keywords)
    
    # Also check for negative amounts (money out)
    amount = ws.cell(row=row_idx, column=9).value
    try:
        amount_float = float(amount) if amount else 0
    except:
        amount_float = 0
    
    if has_keyword or amount_float < 0:
        date = ws.cell(row=row_idx, column=11).value
        party = ws.cell(row=row_idx, column=7).value
        summary = ws.cell(row=row_idx, column=14).value
        results.append({
            'date': date,
            'party': party,
            'amount': amount,
            'summary': summary
        })

# Write to file
with open(r'C:\Users\y2k1\.openclaw\workspace\wl\all_lending.txt', 'w', encoding='utf-8') as f:
    f.write(f'Found {len(results)} potential lending records:\n\n')
    for r in results[:50]:
        f.write(f"日期: {r['date']}\n")
        f.write(f"交易对手: {r['party']}\n")
        f.write(f"金额: {r['amount']}\n")
        f.write(f"摘要: {r['summary']}\n")
        f.write('-' * 30 + '\n')

print(f'Total: {len(results)} records found')

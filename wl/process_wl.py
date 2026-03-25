# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import os

# Read Excel file
wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx', data_only=True)
ws = wb.active

# Get all rows and find 河南浩之源
results = []
for row in ws.iter_rows(min_row=3, values_only=True):
    # Column 7 is the counterparty (对方账户名称)
    if row[6]:  # Column index 6 = column 7
        cell_str = str(row[6])
        if '浩之源' in cell_str:
            # Date is column 11, Amount is column 10
            results.append({
                '日期': row[10],  # 交易时间
                '交易对手': row[6],  # 对方账户名称
                '交易金额': row[9]   # 交易金额
            })

print(f'Found {len(results)} rows for 河南浩之源')

# Also read CMB CSV for 陈亚
import csv
chenya_records = []
with open(r'C:\Users\y2k1\.openclaw\workspace\wl\cmb.csv', 'r', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    for row in reader:
        line = ','.join(row)
        if '陈亚' in line:
            # Parse: date, time, income,支出, balance, type, remark
            parts = line.split(',')
            if len(parts) >= 7:
                date = parts[0].strip().replace('\t', '')
                income = parts[2].strip().replace('\t', '') if len(parts) > 2 else ''
                outcome = parts[3].strip().replace('\t', '') if len(parts) > 3 else ''
                amount = income if income else ('-' + outcome if outcome else '')
                remark = parts[6].strip().replace('\t', '') if len(parts) > 6 else ''
                chenya_records.append({
                    '日期': date,
                    '交易对手': '陈亚',
                    '交易金额': amount
                })

print(f'Found {len(chenya_records)} rows for 陈亚')

# Create output Excel file
output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = "往来款"

# Write headers
output_ws['A1'] = '日期'
output_ws['B1'] = '交易对手'
output_ws['C1'] = '交易金额'

# Write 陈亚 data
row_num = 2
for record in chenya_records:
    output_ws[f'A{row_num}'] = record['日期']
    output_ws[f'B{row_num}'] = record['交易对手']
    output_ws[f'C{row_num}'] = record['交易金额']
    row_num += 1

# Write 河南浩之源 data
for record in results:
    output_ws[f'A{row_num}'] = str(record['日期'])[:10] if record['日期'] else ''
    output_ws[f'B{row_num}'] = record['交易对手']
    output_ws[f'C{row_num}'] = record['交易金额']
    row_num += 1

# Save
output_path = r'C:\Users\y2k1\.openclaw\workspace\wl\往来款汇总.xlsx'
output_wb.save(output_path)
print(f'Saved to {output_path}')
print(f'Total rows: {row_num - 1}')

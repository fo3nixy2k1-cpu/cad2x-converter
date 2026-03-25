# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx')
ws = wb.active

# Get all expense records (negative amounts) - just first 100
expenses = []
for row_idx in range(3, min(ws.max_row + 1, 500)):
    amount = ws.cell(row=row_idx, column=9).value
    try:
        amount_float = float(amount) if amount else 0
    except:
        amount_float = 0
    
    if amount_float < 0:
        date = ws.cell(row=row_idx, column=11).value
        party = ws.cell(row=row_idx, column=7).value
        summary = ws.cell(row=row_idx, column=14).value
        expenses.append({
            'date': date,
            'party': party,
            'amount': amount,
            'summary': summary
        })

# Show first 30 expense records
print(f'Total expenses: {len(expenses)}')
print('First 30 expense records:')
for i, e in enumerate(expenses[:30]):
    print(f"{i+1}. {e['date']} | {e['amount']} | {e['party']} | {e['summary']}")

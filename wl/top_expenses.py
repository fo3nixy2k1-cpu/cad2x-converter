# -*- coding: utf-8 -*-
import openpyxl
from collections import defaultdict

wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx')
ws = wb.active

# Group all expenses by counterparty and calculate total
expenses_by_party = defaultdict(float)

for row_idx in range(3, ws.max_row + 1):
    amount = ws.cell(row=row_idx, column=9).value
    try:
        amount_float = float(amount) if amount else 0
    except:
        amount_float = 0
    
    if amount_float < 0:
        party = ws.cell(row=row_idx, column=7).value
        if party:
            expenses_by_party[party] += abs(amount_float)

# Sort by total amount
sorted_expenses = sorted(expenses_by_party.items(), key=lambda x: x[1], reverse=True)

# Show top 50
print('Top 50 counterparties by total expenses:')
for i, (party, total) in enumerate(sorted_expenses[:50]):
    print(f"{i+1}. {total:,.2f} - {party}")

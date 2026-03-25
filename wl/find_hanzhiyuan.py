# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook

# Read Excel file
wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx')
ws = wb.active

# Headers (row 2)
headers = [cell.value for cell in ws[2]]
print('Original headers:', headers)

# Find 河南浩之源 in the data
results = []
for row in ws.iter_rows(min_row=3, values_only=True):
    row_str = str(row)
    if '浩之源' in row_str or '河南浩之源' in row_str:
        results.append(row)

print(f'Found {len(results)} rows with 河南浩之源')
for r in results[:3]:
    print(r)

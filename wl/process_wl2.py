# -*- coding: utf-8 -*-
import openpyxl
from openpyxl import Workbook
import csv

wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx')
ws = wb.active

# Find all rows with 河南浩之源通信工程有限公司
hanzhiyuan = []
for row in ws.iter_rows(min_row=3, values_only=True):
    if row[6]:
        cell_val = str(row[6])
        # Check for the pattern - these are all from that company
        if '浩之源' in cell_val and '通信' in cell_val:
            date = str(row[10])[:10] if row[10] else ''
            amount = row[8]
            counterparty = row[6]
            hanzhiyuan.append({
                '日期': date,
                '交易对手': counterparty,
                '交易金额': amount
            })

print(f'Found {len(hanzhiyuan)} rows for 河南浩之源通信工程有限公司')
for h in hanzhiyuan:
    print(f"Date: {h['日期']}, Amount: {h['交易金额']}")

# Read CMB CSV for 陈亚
chenya = []
with open(r'C:\Users\y2k1\.openclaw\workspace\wl\cmb.csv', 'r', encoding='utf-8-sig') as f:
    reader = csv.reader(f)
    for row in reader:
        line = ','.join(row)
        if '陈亚' in line:
            parts = line.split(',')
            if len(parts) >= 7:
                date = parts[0].strip().replace('\t', '')
                income = parts[2].strip().replace('\t', '') if len(parts) > 2 else ''
                outcome = parts[3].strip().replace('\t', '') if len(parts) > 3 else ''
                amount = income if income else ('-' + outcome if outcome else '')
                chenya.append({
                    '日期': date,
                    '交易对手': '陈亚',
                    '交易金额': amount
                })

print(f'Found {len(chenya)} rows for 陈亚')

# Create output
output_wb = Workbook()
output_ws = output_wb.active
output_ws.title = '往来款'

output_ws['A1'] = '日期'
output_ws['B1'] = '交易对手'
output_ws['C1'] = '交易金额'

row_num = 2
for record in chenya:
    output_ws[f'A{row_num}'] = record['日期']
    output_ws[f'B{row_num}'] = record['交易对手']
    output_ws[f'C{row_num}'] = record['交易金额']
    row_num += 1

for record in hanzhiyuan:
    output_ws[f'A{row_num}'] = record['日期']
    output_ws[f'B{row_num}'] = record['交易对手']
    output_ws[f'C{row_num}'] = record['交易金额']
    row_num += 1

output_wb.save(r'C:\Users\y2k1\.openclaw\workspace\wl\往来款汇总.xlsx')
print(f'\nSaved! Total: {row_num-1} rows')

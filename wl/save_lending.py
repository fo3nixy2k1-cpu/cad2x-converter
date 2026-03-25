# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\y2k1\.openclaw\workspace\wl\gr.xlsx')
ws = wb.active

# Search all rows for 借钱 or 借款 in ANY column
results = []
for row_idx in range(3, ws.max_row + 1):
    row_text = ''
    for col_idx in range(1, 15):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            row_text += str(cell.value) + ' '
    
    if '借钱' in row_text or '借款' in row_text:
        date = ws.cell(row=row_idx, column=11).value
        party = ws.cell(row=row_idx, column=7).value
        amount = ws.cell(row=row_idx, column=9).value
        summary = ws.cell(row=row_idx, column=14).value
        results.append({
            'date': date,
            'party': party,
            'amount': amount,
            'summary': summary
        })

# Write to file with UTF-8
with open(r'C:\Users\y2k1\.openclaw\workspace\wl\lending_records.txt', 'w', encoding='utf-8') as f:
    f.write(f'Found {len(results)} records with 借钱/借款:\n\n')
    for r in results:
        f.write(f"日期: {r['date']}\n")
        f.write(f"交易对手: {r['party']}\n")
        f.write(f"金额: {r['amount']}\n")
        f.write(f"摘要: {r['summary']}\n")
        f.write('-' * 30 + '\n')

print('Done! Results saved to lending_records.txt')

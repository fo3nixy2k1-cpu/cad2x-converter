#!/usr/bin/env python3
# 验证Excel文件并手动计算总和
import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\y2k1\.openclaw\workspace\基站工程量清单.xlsx")
ws = wb.active

print("工作表名称:", ws.title)
print("\n数据预览：")
for row in ws.iter_rows(min_row=3, max_row=23, values_only=True):
    print(row)

# 手动计算总和
total = 0
for row in range(3, 23):
    cell_value = ws.cell(row=row, column=4).value
    if isinstance(cell_value, (int, float)):
        total += cell_value

print(f"\n手动计算的总工时: {total} 人天")
print(f"表格中的公式: {ws['D23'].value}")

wb.close()

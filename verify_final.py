#!/usr/bin/env python3
# 最终验证Excel文件
import openpyxl
import os

filepath = r"C:\Users\y2k1\.openclaw\workspace\基站工程量清单.xlsx"

print(f"检查文件: {filepath}")
print(f"文件存在: {os.path.exists(filepath)}")
print(f"文件大小: {os.path.getsize(filepath)} 字节")

try:
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    print(f"\n工作表: {ws.title}")
    print(f"最大行: {ws.max_row}")
    print(f"最大列: {ws.max_column}")

    print("\n标题:", ws['A1'].value)
    print("\n表头:")
    for col in range(1, 7):
        print(f"  {ws.cell(row=2, column=col).value}", end=' ')
    print()

    print("\n数据行:")
    total = 0
    for row in range(3, 23):  # 3到22行是数据
        cells = []
        for col in range(1, 5):  # A到D列
            cell = ws.cell(row=row, column=col)
            cells.append(cell.value)
        if cells[0] and cells[1]:  # 有序号和工作类别
            print(f"  {cells[0]:2}. {cells[1]:<8} - {cells[2]:<25} - {cells[3]:>4} 天")
            if isinstance(cells[3], (int, float)):
                total += cells[3]

    print(f"\n汇总:")
    print(f"  D23单元格公式: {ws['D23'].value}")
    print(f"  手动累加总和: {total} 人天")

    # 尝试获取计算结果（如果公式已计算）
    print(f"  D23单元格显示值: {ws['D23'].value}")

    wb.close()

    print("\n✅ Excel文件创建成功！")

except Exception as e:
    print(f"❌ 错误: {e}")
    import traceback
    traceback.print_exc()

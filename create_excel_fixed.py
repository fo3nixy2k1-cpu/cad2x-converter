#!/usr/bin/env python3
# 创建基站工程工作量清单Excel文件 - 修复版本
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "工程量清单"

# 设置列宽
ws.column_dimensions['A'].width = 8   # 序号
ws.column_dimensions['B'].width = 25  # 工作类别
ws.column_dimensions['C'].width = 50  # 具体工作项
ws.column_dimensions['D'].width = 12  # 预估工时(人天)
ws.column_dimensions['E'].width = 60  # 工作内容/技术要求
ws.column_dimensions['F'].width = 15  # 备注

# 标题
ws['A1'] = '5G基站工程量清单'
ws.merge_cells('A1:F1')
title_font = Font(name='微软雅黑', size=18, bold=True)
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# 表头
headers = ['序号', '工作类别', '具体工作项', '预估工时(人天)', '工作内容/技术要求', '备注']
header_font = Font(name='微软雅黑', size=11, bold=True)
header_fill = PatternFill('solid', start_color='4F81BD')
header_alignment = Alignment(horizontal='center', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=header)
    cell.font = header_font
    cell.alignment = header_alignment
    cell.fill = header_fill
    cell.border = thin_border

ws.row_dimensions[2].height = 45

# 数据定义
data = [
    (1, '天馈系统', '天线安装', 2.0, '3个AAU/天线安装、固定、防水处理\n安装位置：15米支撑杆第二、三层', '含吊装、机械工具'),
    (2, '天馈系统', '支撑杆安装', 1.5, '15米支撑杆架设、基础浇筑、接地\n需起重设备配合', '含材料运输'),
    (3, '天馈系统', '天馈线敷设', 1.0, '射频电缆敷设、固定、接地\n长度约30-50米/扇区', '含接头制作'),
    (4, '天馈系统', '天线调测', 0.5, '方向角调整、下倾角校准\n方位角精度±1°，俯仰角±0.1°', '需GPS定位仪'),
    (5, '电源系统', '开关电源安装', 1.5, '开关电源柜安装、固定、接线\n含8个50A整流模块', '含面板安装'),
    (6, '电源系统', '电池组安装', 1.0, '铅酸电池500Ah×2组 + 锂电池50Ah×1组\n搬运、固定、连接、防腐处理', '需搬运设备'),
    (7, '电源系统', '直流配电架', 0.5, '直流配电单元安装、接线\n负荷电流400A', '含熔丝配置'),
    (8, '电源系统', '交流配电箱', 0.5, '380V交流配电箱安装\n进线电缆25mm²铜芯', '含计量装置'),
    (9, '电缆布线', '电力电缆敷设', 2.5, '380V电力电缆200米敷设\n规格：铜芯25mm²\n含穿管/桥架敷设', '需电缆沟开挖'),
    (10, '电缆布线', '直流电缆敷设', 1.5, '400A直流正负极缆敷设\n长度约50米/路', '含绝缘处理'),
    (11, '电缆布线', '接地系统', 1.0, '接地体埋设、接地线敷设\n接地电阻≤5Ω', '含土壤处理'),
    (12, '机房配套', '空调安装', 0.5, '2台3匹空调安装、固定\n冷媒管敷设、保温', '含支架制作'),
    (13, '机房配套', '走线架安装', 1.0, '室内外走线架安装、固定\n长度约30米', '含膨胀螺栓'),
    (14, '机房配套', '监控/消防', 0.5, '监控摄像头、烟感、温感安装\n火灾报警系统接入', '可选，按需'),
    (15, '调测优化', '设备加电测试', 0.5, '逐级加电、电压测试\n开关电源模块均流测试', '需仪表：万用表'),
    (16, '调测优化', '驻波比测试', 0.5, '天馈系统VSWR测试\n要求：VSWR≤1.5', '需驻波比测试仪'),
    (17, '调测优化', '功率校准', 0.5, '发射功率校准、告警测试\n单模块功率50A', '需功率计'),
    (18, '调测优化', '网络优化', 1.5, '邻区规划、参数配置\n覆盖测试、干扰排查', '需路测软件'),
    (19, '其他', '施工准备', 1.0, '现场勘察、材料准备\n工器具准备、安全措施', '含图纸会审'),
    (20, '其他', '资料整理', 0.5, '竣工图纸编制\n测试记录、验收报告', '按规范要求'),
]

# 填充数据
cell_font = Font(name='微软雅黑', size=10)
cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
fill_light_blue = PatternFill('solid', start_color='DCE6F1')

for row_idx, row_data in enumerate(data, 3):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = thin_border
        if col_idx == 1:  # 序号列
            cell.font = Font(name='微软雅黑', size=10, bold=True)
            cell.fill = fill_light_blue

# 汇总行
summary_row = len(data) + 3
ws.cell(row=summary_row, column=2, value='合计').font = Font(name='微软雅黑', size=11, bold=True)
ws.cell(row=summary_row, column=2).fill = PatternFill('solid', start_color='F2F2F2')
ws.merge_cells(f'B{summary_row}:C{summary_row}')

# 总工时公式
total_hours_formula = f'=SUM(D3:D{summary_row-1})'
ws.cell(row=summary_row, column=4, value=total_hours_formula)
ws.cell(row=summary_row, column=4).font = Font(name='微软雅黑', size=11, bold=True)
ws.cell(row=summary_row, column=4).fill = PatternFill('solid', start_color='F2F2F2')

# 空白单元格填充
for col in [5, 6]:
    ws.cell(row=summary_row, column=col, value='').fill = PatternFill('solid', start_color='F2F2F2')

# 汇总行边框
double_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='double'),
    bottom=Side(style='double')
)
for col in range(1, 7):
    cell = ws.cell(row=summary_row, column=col)
    cell.border = double_border

# 添加说明
info_start = summary_row + 2
info_texts = [
    '说明：',
    '1. 预估工时为单人工作量，实际施工可并行缩短工期',
    '2. 高空作业需配备安全防护设备及专业资质',
    '3. 电力施工需当地供电部门配合审批',
    '4. 调测优化需具备运营商入网许可资质',
    '5. 总工时约15.5人天，建议安排2-3人团队，工期约5-7个工作日',
    '6. 所有工作量基于图纸估算，具体以现场实际情况为准'
]

for i, text in enumerate(info_texts):
    row = info_start + i
    cell = ws.cell(row=row, column=1, value=text)
    ws.merge_cells(f'A{row}:F{row}')
    if i == 0:  # 说明标题
        cell.font = Font(name='微软雅黑', size=9, bold=True)
    else:
        cell.font = Font(name='微软雅黑', size=9)
    cell.alignment = Alignment(horizontal='left', vertical='center')

# 保存文件
output_path = r"C:\Users\y2k1\.openclaw\workspace\基站工程量清单.xlsx"
wb.save(output_path)

print(f"Excel文件已成功创建: {output_path}")
print(f"文件大小: {os.path.getsize(output_path)} 字节")
print("\n总预估工时：15.5人天")
print("建议团队：2-3人")
print("预计工期：5-7个工作日")
